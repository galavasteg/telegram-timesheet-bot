import itertools
from collections import defaultdict
from datetime import datetime, timedelta, time
from functools import partial
from io import BytesIO
from math import ceil
from typing import Tuple, List, Generator, Iterable

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from timesheetbot import utils

DEFAULT_TIME_STEP_MINUTES = 30
TIME_INITIAL_ROW = 2
DATE_INITIAL_COL = 2
DATE_COLUMN_WIDTH = len('2021-01-01')
TIME_COLUMN_WIDTH = len('00:00-00:30')
cut_time = partial(datetime.replace, hour=0, minute=0, second=0, microsecond=0)


def generate_report(
    time_range: Tuple[datetime, datetime],
    activities: List[tuple],
    time_step_minutes: int = DEFAULT_TIME_STEP_MINUTES,
) -> BytesIO:
    wb = Workbook(iso_dates=True)
    wb['Sheet'].title = 'Timesheet'
    ws = wb['Timesheet']
    ws.freeze_panes = "B2"
    actvts = prepare_activities(time_range, activities)

    for col, (date_val, activity_gr) in enumerate(
        itertools.groupby(actvts, key=lambda a: cut_time(a[-3])), DATE_INITIAL_COL
    ):
        times = get_report_times(time_step_minutes, init_date_value=date_val)
        # fill date of statistic time frame (1-st row)
        wb['Timesheet'].cell(column=col, row=1, value=str(date_val.date()))
        ws.column_dimensions[get_column_letter(col)].width = DATE_COLUMN_WIDTH

        activities = tuple(activity_gr)
        for row, time_frame in enumerate(zip(times, times[1:]), TIME_INITIAL_ROW):
            if col == DATE_INITIAL_COL:
                # fill time frames (1-st column)
                ws.cell(column=1, row=row, value='-'.join(map(lambda t: datetime.strftime(t, '%H:%M'), time_frame)))
                ws.column_dimensions[get_column_letter(1)].width = TIME_COLUMN_WIDTH

            category_cell = ws.cell(column=col, row=row)
            if time_frame[1] < time_range[0] or time_range[1] < time_frame[0]:
                # set color for cells out of time_range
                category_cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

            # fill in activities
            category_cell.value = get_longest_frame_category(activities, time_frame)

    return _report_to_virt_file(wb)


def get_longest_frame_category(activities: Iterable, time_frame: Tuple[datetime, datetime]) -> str:
    t0, t1 = time_frame
    get_category_name = lambda activity: activity[-1]
    get_frame_activity = lambda activity: any(t0 <= a_t <= t1 for a_t in activity[-3:-1])

    def iter_category_duration() -> Generator[Tuple[str, timedelta], None, None]:
        for activity in activities:
            if not (category := get_category_name(activity)) or not get_frame_activity(activity):
                continue

            if activity[-3] < t0 < activity[-2]:
                yield category, activity[-2] - t0
            elif all(t0 <= a_t <= t1 for a_t in activity[-3:-1]):
                yield category, activity[-2] - activity[-3]
            elif activity[-3] < t1 <= activity[-2]:
                yield category, t1 - activity[-3]
                break

    category_frame_duration = defaultdict(lambda: timedelta(0))
    for cat, dur in iter_category_duration():
        category_frame_duration[cat] += dur

    return max(category_frame_duration, key=category_frame_duration.get, default='')


def prepare_activities(time_range, activities) -> tuple:
    # statistic time frame
    dates = tuple((datetime.combine(date_val, time()), None, None) for date_val in get_report_dates(time_range))

    def prepare_activity(activity: tuple) -> tuple:
        return *activity[:-3], *map(utils.parse_datetime, activity[-3:-1]), activity[-1]
    # merge time frame with activities
    activities = sorted(dates + tuple(map(prepare_activity, activities)), key=lambda a: a[-3].date())

    return tuple(activities)


def get_report_dates(time_range: Tuple[datetime, datetime]) -> Generator[datetime, None, None]:
    assert time_range[0] < time_range[1]
    d0, d1 = map(cut_time, time_range)
    days_cnt = ceil((d1 - d0) / timedelta(1)) + 1
    for delta_days in range(days_cnt):
        yield d0 + timedelta(delta_days)


def get_report_times(time_step_minutes: int, init_date_value: datetime) -> Tuple[datetime, ...]:
    report_row_cnt = ceil(timedelta(1) / timedelta(minutes=time_step_minutes))
    time_step = timedelta(minutes=time_step_minutes)
    times = tuple(itertools.accumulate(range(report_row_cnt), lambda tm, _: tm + time_step, initial=init_date_value))
    return times


def _report_to_virt_file(wb: Workbook):
    virtual_wb = BytesIO()
    wb.save(virtual_wb)
    virtual_wb.seek(0)
    return virtual_wb
